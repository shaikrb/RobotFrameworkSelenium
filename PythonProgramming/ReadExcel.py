import openpyxl

#Load workbook
wb = openpyxl.load_workbook("..//resources//TestExcel.xlsx")

#Print all sheet names
print(wb.sheetnames)

#Get active sheet
print("Active Sheet: "+ wb.active.title)

#Load sheet
sheet = wb['Hello']
print(sheet.title)

#Get Cell content
print(sheet['B2'].value)

#Other approach
cell1 = sheet.cell(2,3)
print(cell1.value)

for i in range (1,sheet.max_row+1):
    for j in range (1,sheet.max_column+1):
        c=sheet.cell(i,j)
        print(c.value)

for r in sheet['A1':'C2']:
    for c in r:
        print(c.value)